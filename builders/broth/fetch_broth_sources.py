#!/usr/bin/env python3
"""fetch_broth_sources.py — pull the two broth source Sheets as CSV, with NO browser and NO Claude.

Auth: the same Google service account the daily tab uses (env GOOGLE_SA_JSON = the key file's JSON
text), which both Sheets are shared with as Viewer.

Route: Drive v3 files.export to **xlsx**, then openpyxl. files.export to text/csv only ever returns
the FIRST tab, and the Mapal broker Sheet keeps its cells and its deviations on different tabs, so
csv is not an option here. Reference: reference_sources_and_routes, "Drive xlsx + openpyxl".

  python3 fetch_broth_sources.py --out /tmp/broth_src
      -> broth_cells.csv  broth_deviations.csv  refractometer.csv

Guards (loud, non-zero exit): auth failure, HTTP != 200, an HTML body (a wrong Google id fails
SILENTLY at 200 with a "Page not found" page), a missing tab, or a tab with no data rows.
"""
import argparse, csv, datetime, io, json, os, re, sys
import requests
from google.oauth2 import service_account
from google.auth.transport.requests import Request
from openpyxl import load_workbook

BROKER = "17M3-5veLg5rnLlCOvCSjBspzzWc8GIGCBPK_M5yzMrw"   # "Mapal Broker Data", michael@makiramen.com
FACTORY = "1_xRTe_AaPOsfhpP9mtrZYBvOgxsh_1daCUqJFSLTtBA"   # "Refractometer Reading (Responses)"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def log(*a): print(*a, file=sys.stderr)

def token():
    raw = os.environ.get("GOOGLE_SA_JSON")
    if not raw:
        log("GOOGLE_SA_JSON is not set"); sys.exit(2)
    creds = service_account.Credentials.from_service_account_info(
        json.loads(raw), scopes=["https://www.googleapis.com/auth/drive.readonly"])
    creds.refresh(Request())
    return creds.token

def export_xlsx(tok, fid, label):
    r = requests.get("https://www.googleapis.com/drive/v3/files/%s/export" % fid,
                     params={"mimeType": XLSX},
                     headers={"Authorization": "Bearer %s" % tok}, timeout=180)
    if r.status_code != 200:
        log("export %s (%s) -> HTTP %d: %s" % (label, fid, r.status_code, r.text[:300])); sys.exit(3)
    b = r.content
    if b[:2] != b"PK":
        log("export %s did not return an xlsx (first bytes %r) — wrong id or no access" % (label, b[:20])); sys.exit(3)
    log("%s: %d B xlsx" % (label, len(b)))
    return load_workbook(io.BytesIO(b), read_only=True, data_only=True)

def clean(c):
    """Two traps this exists for, both found by the first live run on 03/09/2026.
    1. MAPAL PUTS NON-BREAKING SPACES IN LOCATION NAMES. "Maki\\xa0O2\\xa0Arena" is not
       "Maki O2 Arena", so it reads as a brand new site and silently forks its own series.
    2. openpyxl HANDS BACK REAL date OBJECTS for a date-formatted cell, so str() gives
       "2026-08-31 00:00:00" and every dd/mm/yyyy parser downstream returns None. That blanked
       the whole factory line and the bake died on an empty mean."""
    if c is None:
        return ""
    if isinstance(c, datetime.datetime):
        return c.strftime("%Y-%m-%d %H:%M:%S") if (c.hour or c.minute or c.second) else c.strftime("%Y-%m-%d")
    if isinstance(c, datetime.date):
        return c.isoformat()
    return " ".join(str(c).replace("\u00a0", " ").split())


def dump(wb, sheet, path, label, required=True):
    if sheet not in wb.sheetnames:
        if not required:
            log("%s: tab %r absent, skipped" % (label, sheet)); return 0
        log("%s: tab %r is not in %s" % (label, sheet, wb.sheetnames)); sys.exit(4)
    ws = wb[sheet]
    rows = 0
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for r in ws.iter_rows(values_only=True):
            if all(clean(c) == "" for c in r):
                continue
            w.writerow([clean(c) for c in r])
            rows += 1
    log("%s/%s: %d rows -> %s" % (label, sheet, rows, os.path.basename(path)))
    return rows

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    a = ap.parse_args()
    os.makedirs(a.out, exist_ok=True)
    tok = token()

    wb = export_xlsx(tok, BROKER, "broker")
    n = dump(wb, "broth_cells", os.path.join(a.out, "broth_cells.csv"), "broker")
    if n < 2:
        log("broker/broth_cells has no data rows — the Apps Script pull is dead, stop here"); sys.exit(5)
    # deviations is fail-soft: the Mapal endpoint has been 504ing since 01/09/2026 and the rows it
    # does return carry a null location, so the tab can legitimately be empty or stale.
    if not dump(wb, "broth_deviations", os.path.join(a.out, "broth_deviations.csv"), "broker", required=False):
        open(os.path.join(a.out, "broth_deviations.csv"), "w", encoding="utf-8").write(
            "deviation_id,location,kind,date,type,state,open_closed,action,closed_date\n")

    fwb = export_xlsx(tok, FACTORY, "factory")
    n = dump(fwb, fwb.sheetnames[0], os.path.join(a.out, "refractometer.csv"), "factory")
    if n < 2:
        log("the refractometer sheet has no data rows"); sys.exit(5)
    print("ok")

if __name__ == "__main__":
    main()
